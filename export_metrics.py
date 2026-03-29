"""
Export Forecast Metrics to Excel
Writes a multi-sheet report with live Excel formulas and
professional formatting using openpyxl.

Sheets
------
1. Raw Data       – cleaned data plus formula columns (Error, |Error|, APE)
2. By Store_Dept  – grouped metrics with formulas referencing Raw Data
3. Overall        – summary metrics with formulas referencing Raw Data
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from ingestion_etl import ingest_and_etl

# ─── Constants ────────────────────────────────────────────────────────────────
OUTPUT_FILE = "forecast_metrics_report.xlsx"
SOURCE_FILE = "sales_forecast_data.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Number format strings
FMT_CURRENCY = '#,##0.00'
FMT_INTEGER = '0'
FMT_DECIMAL2 = '0.00'
FMT_PERCENT = '0.00"%"'
FMT_RATIO = '0.0000'


# ─── Helpers ──────────────────────────────────────────────────────────────────
def _style_header(ws):
    """Apply header styling: font, fill, alignment, border."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _autofit_columns(ws, min_width=10, padding=3):
    """Auto-fit column widths based on cell content."""
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            max_len + padding, min_width
        )


def _apply_border(ws):
    """Apply thin borders to all data cells."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER


def _freeze_top_row(ws):
    """Freeze the header row."""
    ws.freeze_panes = "A2"


# ─── Sheet builders ───────────────────────────────────────────────────────────
def _write_raw_data(df, writer):
    """
    Write raw data to the 'Raw Data' sheet with formula columns:
      Error          = weekly_sales − forecast   (col H)
      Absolute Error = ABS(Error)                (col I)
      APE            = |Error| / |weekly_sales|  (col J, as percentage)
    """
    # Write the base data (cols A–F: store, dept, year, week, forecast, weekly_sales)
    # Reorder so forecast (E) comes before weekly_sales (F) for natural reading
    base = df[["store", "dept", "year", "week", "forecast", "weekly_sales"]].copy()
    base.to_excel(writer, sheet_name="Raw Data", index=False)

    ws = writer.sheets["Raw Data"]
    last_row = ws.max_row  # includes header

    # Add formula column headers
    ws["G1"] = "Error"
    ws["H1"] = "Absolute_Error"
    ws["I1"] = "APE"

    # Write formulas for each data row
    # Columns: E=forecast, F=weekly_sales, G=error, H=abs_error, I=APE
    for r in range(2, last_row + 1):
        ws[f"G{r}"] = f"=F{r}-E{r}"                        # Error
        ws[f"H{r}"] = f"=ABS(G{r})"                        # |Error|
        ws[f"I{r}"] = f'=IF(F{r}=0,"N/A",H{r}/ABS(F{r}))' # APE

    # Formatting
    _style_header(ws)
    _freeze_top_row(ws)

    for r in range(2, last_row + 1):
        for col in ["A", "B", "C", "D"]:       # store, dept, year, week
            ws[f"{col}{r}"].number_format = FMT_INTEGER
        for col in ["E", "F", "G"]:             # forecast, weekly_sales, error
            ws[f"{col}{r}"].number_format = FMT_CURRENCY
        ws[f"H{r}"].number_format = FMT_CURRENCY  # |error|
        ws[f"I{r}"].number_format = FMT_DECIMAL2   # APE

    _apply_border(ws)
    _autofit_columns(ws)

    return last_row  # needed by other sheets for formula ranges


def _write_grouped_metrics(df, writer, raw_last_row):
    """
    Write the 'By Store_Dept' sheet with formulas.

    For each unique (store, dept) pair, writes SUMPRODUCT / COUNTIFS
    formulas that reference the Raw Data sheet.
    """
    # Get unique (store, dept) pairs preserving order
    pairs = df[["store", "dept"]].drop_duplicates().reset_index(drop=True)

    ws = writer.book.create_sheet("By Store_Dept")

    # Headers
    headers = ["Store", "Dept", "MAD", "MAPE", "Sales_to_Forecast_Ratio",
               "Forecast_Bias"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    rd = "'Raw Data'"  # sheet reference for formulas
    lr = raw_last_row  # last data row

    def rng(col):
        """Return absolute range string like $A$2:$A$191."""
        return f"${col}$2:${col}${lr}"

    for i, (_, row) in enumerate(pairs.iterrows(), start=2):
        store = int(row["store"])
        dept = int(row["dept"])

        ws.cell(row=i, column=1, value=store)
        ws.cell(row=i, column=2, value=dept)

        # Criteria for SUMPRODUCT matching: (store match)*(dept match)
        match = (
            f"({rd}!{rng('A')}={store})*({rd}!{rng('B')}={dept})"
            f"*({rd}!{rng('F')}<>0)"
        )

        # MAD = mean of |error| for this group
        ws.cell(row=i, column=3).value = (
            f"=SUMPRODUCT({match}*{rd}!{rng('H')})"
            f"/SUMPRODUCT({match}*1)"
        )

        # MAPE = mean of APE × 100
        ws.cell(row=i, column=4).value = (
            f"=SUMPRODUCT({match}*{rd}!{rng('I')})"
            f"/SUMPRODUCT({match}*1)*100"
        )

        # Sales-to-Forecast Ratio = Σ sales / Σ forecast
        ws.cell(row=i, column=5).value = (
            f"=SUMPRODUCT({match}*{rd}!{rng('F')})"
            f"/SUMPRODUCT({match}*{rd}!{rng('E')})"
        )

        # Forecast Bias = mean of error
        ws.cell(row=i, column=6).value = (
            f"=SUMPRODUCT({match}*{rd}!{rng('G')})"
            f"/SUMPRODUCT({match}*1)"
        )

    # Formatting
    _style_header(ws)
    _freeze_top_row(ws)

    last = len(pairs) + 1
    for r in range(2, last + 1):
        ws.cell(row=r, column=1).number_format = FMT_INTEGER
        ws.cell(row=r, column=2).number_format = FMT_INTEGER
        ws.cell(row=r, column=3).number_format = FMT_CURRENCY   # MAD
        ws.cell(row=r, column=4).number_format = FMT_DECIMAL2   # MAPE
        ws.cell(row=r, column=5).number_format = FMT_RATIO      # Ratio
        ws.cell(row=r, column=6).number_format = FMT_CURRENCY   # Bias

    _apply_border(ws)
    _autofit_columns(ws)


def _write_overall(writer, raw_last_row):
    """
    Write the 'Overall' sheet with formulas referencing Raw Data.
    """
    ws = writer.book.create_sheet("Overall")

    headers = ["Metric", "Value"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    rd = "'Raw Data'"
    lr = raw_last_row

    def rng(col):
        """Return absolute range string like $A$2:$A$191."""
        return f"${col}$2:${col}${lr}"

    nz = f"({rd}!{rng('F')}<>0)"  # non-zero sales filter

    metrics = [
        ("MAD",
         f"=SUMPRODUCT({nz}*{rd}!{rng('H')})/SUMPRODUCT({nz}*1)"),
        ("MAPE",
         f"=SUMPRODUCT({nz}*{rd}!{rng('I')})/SUMPRODUCT({nz}*1)*100"),
        ("Sales-to-Forecast Ratio",
         f"=SUMPRODUCT({nz}*{rd}!{rng('F')})/SUMPRODUCT({nz}*{rd}!{rng('E')})"),
        ("Forecast Bias",
         f"=SUMPRODUCT({nz}*{rd}!{rng('G')})/SUMPRODUCT({nz}*1)"),
    ]

    fmt_map = {
        "MAD": FMT_CURRENCY,
        "MAPE": FMT_DECIMAL2,
        "Sales-to-Forecast Ratio": FMT_RATIO,
        "Forecast Bias": FMT_CURRENCY,
    }

    for r, (name, formula) in enumerate(metrics, start=2):
        ws.cell(row=r, column=1, value=name).font = Font(bold=True)
        ws.cell(row=r, column=2).value = formula
        ws.cell(row=r, column=2).number_format = fmt_map[name]

    # Extra context rows
    ws.cell(row=7, column=1, value="Total Records").font = Font(bold=True)
    ws.cell(row=7, column=2).value = f"=COUNTA({rd}!{rng('A')})"
    ws.cell(row=7, column=2).number_format = "#,##0"

    ws.cell(row=8, column=1, value="Records Used (non-zero sales)").font = Font(bold=True)
    ws.cell(row=8, column=2).value = f"=SUMPRODUCT({nz}*1)"
    ws.cell(row=8, column=2).number_format = "#,##0"

    # Formatting
    _style_header(ws)
    _freeze_top_row(ws)
    _apply_border(ws)
    _autofit_columns(ws, min_width=14)


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    """Build the forecast metrics Excel report."""
    df = ingest_and_etl(SOURCE_FILE)
    if df is None or df.empty:
        print("❌ Failed to load data. Exiting.")
        return

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        # 1. Raw Data (must be first — other sheets reference it)
        raw_last_row = _write_raw_data(df, writer)

        # 2. Grouped metrics with formulas
        _write_grouped_metrics(df, writer, raw_last_row)

        # 3. Overall summary with formulas
        _write_overall(writer, raw_last_row)

    # Reorder sheets: By Store_Dept, Overall, Raw Data
    wb = load_workbook(OUTPUT_FILE)
    order = ["By Store_Dept", "Overall", "Raw Data"]
    wb._sheets = [wb[name] for name in order]
    wb.save(OUTPUT_FILE)

    print(f"✅ Report saved to '{OUTPUT_FILE}'")
    print(f"   Sheets: {' | '.join(order)}")
    print(f"   Raw data rows: {len(df)}")


if __name__ == "__main__":
    main()
